import tkinter as tk
from tkinter import ttk
import sys
import plotly.graph_objects as go
import plotly.express as px
import math
from openpyxl.styles import Font, Color
from openpyxl import Workbook

### Tk App ###

class App:
    def __init__(self, root, options, df, specifications):
        self.root = root
        self.options = options
        self.selected_options = []  # Keep track of selected options
        self.df = df
        self.specifications = specifications

        self.create_widgets()

    def create_widgets(self):

        # Create a Treeview widget
        self.treeview = ttk.Treeview(self.root, columns=("Variable", "Min", "Max", "Value", "Type", "Input/Output"), show="headings")
        self.treeview.pack(padx=10, pady=5)

        # Add column headings
        self.treeview.heading("Variable", text="Variable")
        self.treeview.heading("Min", text="Min")
        self.treeview.heading("Max", text="Max")
        self.treeview.heading("Value", text="Value")
        self.treeview.heading("Type", text="Type")
        self.treeview.heading("Input/Output", text="Input/Output")

        # Define color tags based on type and min/max columns
        type_colors = {
            #"bounds": "black",
            "ineq_cstr": "grey",
            "objective": "red",
            "eq_cstr": "green",
            #"free": "black",
        }

        type_names = {
            "bounds": "constrained",
            "ineq_cstr": "inequality",
            "objective": "objective",
            "eq_cstr": "fixed",
            "free": "free",
        }

        # Add options and additional information
        for option in self.options:
            type_ = self.specifications.Type[option]
            value = self.specifications.Value[option]
            input_output = self.specifications.In_Out[option]

            if isinstance(value, list) and type_=="bounds":
                min_value, max_value = value
                value_text = ""

            elif isinstance(value, list) and type_=="ineq_cstr":
                min_value, max_value = "", ""
                value_text = str(value[0]) + "      " + str(value[1])

            elif (isinstance(value, int) or isinstance(value, float)) and not math.isnan(value):
                min_value = ""
                max_value = ""
                value_text = value

            else:
                min_value = ""
                max_value = ""
                value_text = ""
 
            # Add rows and colors
            self.treeview.insert("", "end", values=(option, min_value, max_value, value_text, type_names.get(type_), input_output), tags=(type_,))
            self.treeview.tag_configure(type_, foreground=type_colors.get(type_, "black"))

        # Create a scrollbar for the Treeview
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.treeview.yview)
        self.treeview.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        # Create a context menu for the Treeview
        self.option_menu = tk.Menu(self.root, tearoff=False)
        self.option_menu.add_command(label="Additional Info", command=self.show_additional_info)
        self.option_menu.entryconfigure(0, state="disabled")  # Disable the "Additional Info" menu item initially
        self.treeview.bind("<Button-3>", self.show_context_menu)

        self.export_button = tk.Button(self.root, text="Export", command=self.export_to_excel)
        self.export_button.pack(pady=5)

        self.exit_button = tk.Button(self.root, text="Exit", command=self.exit_program)
        self.exit_button.pack(pady=5)


    def show_context_menu(self, event):
        selection = self.treeview.selection()
        if selection:
            self.selected_options = [self.options[self.treeview.index(item)] for item in selection]  # Update selected options
            menu_label = f"Additional Info ({len(self.selected_options)} options selected)"
            self.option_menu.entryconfigure(0, label=menu_label)
            self.option_menu.entryconfigure(0, state="normal")  # Enable the "Additional Info" menu item
            if len(self.selected_options) > 1:
                self.option_menu.entryconfigure(1, label="Plot", command=self.plot_multiple_selected_options)
            else:
                self.option_menu.entryconfigure(1, label="Plot", command=self.plot_one_selected_option)
            self.option_menu.tk_popup(event.x_root, event.y_root)


    def show_additional_info(self):
        for selected_option in self.selected_options:
            print(f"Additional info for {selected_option}")

    def export_to_excel(self):
        # Create a Workbook object
        workbook = Workbook()

        # Create a new sheet
        sheet = workbook.active

        # Get the columns to export excluding 'IsBestSolution', 'IsSolution', and 'pareto_pts'
        export_columns = [col for col in self.df.columns if col not in ['IsBestSolution', 'IsSolution', 'pareto_pts']]

        # Write the column names to the sheet
        sheet.append(export_columns)

        # Write the row values to the sheet
        for _, row in self.df[export_columns].iterrows():
            values = row.tolist()
            sheet.append(values)   

        # Apply conditional text color to the desired columns
        for column in self.specifications.index:
            current_spec = self.specifications.Type[column]
            if current_spec == "bounds":
                spec_values = self.specifications.Value[column]
                min_value, max_value = spec_values[0], spec_values[1]
                font_color = "000000"  # Default black color
                for index, value in self.df[column].items():
                    if value <= min_value:
                        font_color = "0000FF"  # Blue color
                    elif value >= max_value:
                        font_color = "FF0000"  # Red color
                    if column in export_columns:
                        cell = sheet.cell(row=index+2, column=export_columns.index(column)+1)
                        cell.font = Font(color=Color(rgb=font_color))
                        cell.value = value
            elif current_spec == "eq_cstr":
                font_color = "808080"  # Grey color
                for index in self.df.index:
                    if column in export_columns:
                        cell = sheet.cell(row=index+2, column=export_columns.index(column)+1)
                        cell.font = Font(color=Color(rgb=font_color))
                        cell.value = self.df[column][index]
            elif current_spec == "objective":
                font_color = "00FF00"  # Green color
                for index in self.df.index:
                    if column in export_columns:
                        cell = sheet.cell(row=index+2, column=export_columns.index(column)+1)
                        cell.font = Font(color=Color(rgb=font_color))
                        cell.value = self.df[column][index]

        # Save the workbook to a file
        workbook.save("exported_data.xlsx")
        print("Result saved!")

    def plot_multiple_selected_options(self):

        fig = go.Figure()
        fig_title = "1d-plot: "

        for selected_option in self.selected_options:
            fig_title += (str(selected_option) + " ")
            fig.add_trace(go.Scatter(x=self.df['IterationNumber'], 
                                     y=self.df[selected_option], 
                                     mode='lines+markers', 
                                     name=selected_option,
                                     hovertemplate=f"{'IterationNumber'}=%{{x}}<br>{selected_option}=%{{y:.2f}}<extra></extra>"))

            if len(self.selected_options) <= 3:
                # Label each dot
                for iteration, x, y in zip(self.df['IterationNumber'], self.df['IterationNumber'], self.df[selected_option]):
                    fig.add_annotation(x=x, y=y, text=str(iteration), showarrow=True, arrowhead=1, ax=0, ay=-20)

        fig.update_layout(
            title={
            'text': fig_title,
            'x': 0.5,
            'xanchor': 'center',
            'yanchor': 'top'
            },
            xaxis_title='Iteration Number',
            yaxis_title='Option Values',
            xaxis=dict(tickmode='linear'),
            yaxis=dict(tickformat='g'),
            legend=dict(x=1, y=1)
        )

        config = {
            'modeBarButtonsToAdd': [
                'downloadImage'
            ]
        }
        fig.show(config=config)

    def plot_one_selected_option(self):
        selected_option = self.selected_options[0]

        color = 'midnightblue'  # Set the desired color value

        fig = px.line(
            self.df,
            x='IterationNumber',
            y=selected_option,
            title="1d-plot: " + str(selected_option),
            labels={'IterationNumber': 'Iteration Number', selected_option: selected_option},
            hover_data={'IterationNumber': True, selected_option: ':.2f'},
            color_discrete_sequence=[color]
        )

        fig.add_trace(
            go.Scatter(
                x=self.df['IterationNumber'],
                y=self.df[selected_option],
                mode='markers',
                name=selected_option,
                hovertemplate=f"{'IterationNumber'}=%{{x}}<br>{selected_option}=%{{y:.2f}}<extra></extra>",
                marker=dict(color=color)
            )
        )

        # Label each dot
        for iteration, x, y in zip(self.df['IterationNumber'], self.df['IterationNumber'], self.df[selected_option]):
            fig.add_annotation(x=x, y=y, text=str(iteration), showarrow=True, arrowhead=1, ax=0, ay=-20)

        current_spec = self.specifications.Value[selected_option]
        current_type = self.specifications.Type[selected_option]

        if isinstance(current_spec, list):
            for i in current_spec:
                if i == current_spec[0]:
                    color = 'blue' if current_type == 'bounds' else 'grey'  # Set color for the first line
                else:
                    color = 'red' if current_type == 'bounds' else 'grey'  # Set color for the second line

                fig.add_shape(
                    type="line",
                    x0=self.df['IterationNumber'].min(),
                    y0=i,
                    x1=self.df['IterationNumber'].max(),
                    y1=i,
                    line=dict(color=color, dash='dash'),
                    name=current_type
                )

        elif isinstance(current_spec, (int, float)) and not math.isnan(current_spec):
            color = 'black' if current_type == 'eq_cstr' else 'red'  # Set color based on type
            fig.add_shape(
                type="line",
                x0=self.df['IterationNumber'].min(),
                y0=current_spec,
                x1=self.df['IterationNumber'].max(),
                y1=current_spec,
                line=dict(color=color, dash='dash'),
                name=current_type
            )


        fig.update_layout(
            title={
            'text': "1d-plot: " + str(selected_option),
            'x': 0.5,
            'xanchor': 'center',
            'yanchor': 'top'
            },
            xaxis_title='Iteration Number',
            yaxis_title=selected_option,
            xaxis=dict(tickmode='linear'),
            yaxis=dict(tickformat='g'),
            legend=dict(x=1, y=1)
        )
        
        config = {
            'modeBarButtonsToAdd': [
                'downloadImage'
            ]
        }
        fig.show(config=config)

    def exit_program(self):
        sys.exit()
